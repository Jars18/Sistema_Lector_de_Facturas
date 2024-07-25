#SISTEMA LECTOR DE FACTURAS
import os
import fitz
from ultralytics import YOLO
import glob
import re
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
from PIL import Image
import easyocr
import cv2
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule
import torch
import sys

import tkinter as tk
from tkinter import filedialog
import shutil
from PIL import Image, ImageTk

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

########################################## SECCION 1: CREACION O APERTURA DEL LIBRO DE TRABAJO ##########################################
if getattr(sys, 'frozen', False):       # Obtener el directorio del ejecutable
    application_path = os.path.dirname(sys.executable)
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

file_path = os.path.join(application_path, "Informacion.xlsx")  # Definir la ruta completa del archivo de Excel
                         
if os.path.exists(file_path):                       #Si ya existe la ruta
    libro = openpyxl.load_workbook(file_path)       #carga el libro
    hoja = libro.active                             #para usarlo
else:
    libro = Workbook()                              #Si no existe lo crea con 
    hoja = libro.active                             #las siguiente caracteristicas:

    #Encabezados
    encabezados = [
        "factura", "Fecha", "Senor", "NIT-CI", "Consumidor", "val_consumidor", "Medidor",
        "Direccion", "Ciudad o Localidad", "Actividad", "Carta Factura", "Remesa y Ruta", "Mes",
        "Categoria Tarifaria", "Fecha de Lectura Anterior", "Fecha de Lectura Actual",
        "Lectura Medidor Anterior", "Lectura Medidor Actual", "Bloque Alto Anterior",
        "Bloque Alto Actual", "Bloque Alto kWh", "Bloque Medio Anterior", "Bloque Medio Actual",
        "Bloque Medio kWh", "Bloque Bajo Anterior", "Bloque Bajo Actual", "Bloque Bajo kWh",
        "Tipo de Lectura", "Multiplicador", "Energia consumida en -n- dias", "Energia Estimada",
        "Energia adicional por cambio de medidor", "Menos devolucion kWh", "Perdidas en el transformador",
        "Total energia a facturar", "Total energia a facturar valid", "Potencia Contratada",
        "Potencia Leida Bloque Alto", "Potencia Leida Bloque Medio", "Potencia Leida Bloque Bajo",
        "Potencia a Facturar", "Exceso de potencia fuera de punta", "Energia Reactiva", "Factor de potencia",
        "Importe por Cargo Fijo", "Importe por Energia", "Importe por Energia - Bloque Alto",
        "Importe por Energia - Bloque Medio", "Importe por Energia - Bloque Bajo", "Importe por Potencia",
        "Importe por Exceso de potencia fuera de punta", "Importe por Consumo", "Importe total por consumo",
        "Importe por bajo factor de potencia", "Mas intereses por Mora", "Mas cargo por conexion",
        "Mas debito consumo no facturado", "Menos credito por devolucion", "Importe total por el suministro",
        "Tasas por Alumbrado Publico", "Tasas por Aseo Urbano", "Importe total factura", "Mas deposito de garantia",
        "mas credito aplicado en cuenta corriente", "menos credito anterior", "Menos reduccion calidad servicio tecnico",
        "Pagos adelantados aplicados", "Importe del mes a cancelar", "Mas deudas pendientes de energia",
        "Deudas pendientes de tasa de aseo", "Menos pago adelantado", "Pago adelantado a la fecha",
        "Importe total a cancelar", "Resumen mes y ano", "Importe"
    ]

    for col_num, header in enumerate(encabezados, 1):   #Inserta los encabezados
        hoja.cell(row=1, column=col_num, value=header)  #en la primera fila

    # Definir el formato condicional (REGLAS DE EXCEL PARA VALIDACION DE DATOS)
    # En esta secion se define el formato de relleno, color de texto y color del borde
    red_fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
    red_font = Font(color="9c0006")
    red_border = Border(
        left=Side(style='thin', color='9c0006'),
        right=Side(style='thin', color='9c0006'),
        top=Side(style='thin', color='9c0006'),
        bottom=Side(style='thin', color='9c0006')
    )

    blue_fill = PatternFill(start_color="b8cce4", end_color="b8cce4", fill_type="solid")
    blue_font = Font(color="1f497d")
    blue_border = Border(
        left=Side(style='thin', color='1f497d'),
        right=Side(style='thin', color='1f497d'),
        top=Side(style='thin', color='1f497d'),
        bottom=Side(style='thin', color='1f497d')
    )

    lightgray_fill = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")

    yellow_fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
    yellow_font = Font(color="9c5700")
    yellow_border = Border(
        left=Side(style='thin', color='9c5700'),
        right=Side(style='thin', color='9c5700'),
        top=Side(style='thin', color='9c5700'),
        bottom=Side(style='thin', color='9c5700')
    )

    purple_fill = PatternFill(start_color="ccc0da", end_color="ccc0da", fill_type="solid")
    purple_font = Font(color="60497a")
    purple_border = Border(
        left=Side(style='thin', color='60497a'),
        right=Side(style='thin', color='60497a'),
        top=Side(style='thin', color='60497a'),
        bottom=Side(style='thin', color='60497a')
    )

    brown_fill = PatternFill(start_color="c4bd97", end_color="c4bd97", fill_type="solid")
    brown_font = Font(color="ab7942")
    brown_border = Border(
        left=Side(style='thin', color='ab7942'),
        right=Side(style='thin', color='ab7942'),
        top=Side(style='thin', color='ab7942'),
        bottom=Side(style='thin', color='ab7942')
    )

    orange_fill = PatternFill(start_color="fcd5b4", end_color="fcd5b4", fill_type="solid")
    orange_font = Font(color="e26b0a")
    orange_border = Border(
        left=Side(style='thin', color='e26b0a'),
        right=Side(style='thin', color='e26b0a'),
        top=Side(style='thin', color='e26b0a'),
        bottom=Side(style='thin', color='e26b0a')
    )

    green_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
    green_font = Font(color="006100")
    green_border = Border(
        left=Side(style='thin', color='006100'),
        right=Side(style='thin', color='006100'),
        top=Side(style='thin', color='006100'),
        bottom=Side(style='thin', color='006100')
    )

    #Definir las fórmulas y rangos
    #-ESTAS REGLAS SON PARA LOS DATOS O CAMPOS QUE NO FUERON RECONOCIDOS Y DEBIERON SERLO:
    #--DE FORMA GENERAL
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(B2))', "B2:I1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(L2))', "L2:P1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(AB2))', "AB2:AB1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(AD2))', "AD2:AD1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(AI2))', "AI2:AJ1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(AZ2))', "AZ2:BA1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(BG2))', "BG2:BG1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(BJ2))', "BJ2:BJ1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(BP2))', "BP2:BP1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(BU2))', "BU2:BW1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #--CON CONDICIONANTE: SI LECTURA MEDIDOR ANTERIOR FUE RECONOCIDO TAMBIEN DEBE SER RECONOCIDO LECTURA MEDIDOR ACTUAL Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($Q2)), ISBLANK($R2))', "R2:R1048576"),
        ('=AND(NOT(ISBLANK($R2)), ISBLANK($Q2))', "Q2:Q1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #--CON CONDICIONANTE: SI UNA LECTURA DE BLOQUE FUE IDENTIFICADO TODOS LOS DEMAS CAMPOS DE LECTURA DE BLOQUE TAMBIEN DEBERIAN SER IDENTIFICADOS
    formula2 = '=AND(SUM(--(NOT(ISBLANK($T2:$W2)))) > 0, ISBLANK(S2))'
    rango2 = "S2:AA1048576"
    rule2 = FormulaRule(formula=[formula2], fill=red_fill, font=red_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #--CON CONDICIONANTE: SI UNA POTENCIA LEIDA DE BLOQUE BAJO O MEDIO FUE IDENTIFICADA TODOS LOS CAMPOS DE POTENCIA LEIDA DEBEN SER RECONOCIDO
    formula2 = '=AND(OR(NOT(ISBLANK($AN2)), NOT(ISBLANK($AM2))), OR(ISBLANK($AM2), ISBLANK($AL2), ISBLANK($AN2)))'
    rango2 = "AL2:AN1048576"
    rule2 = FormulaRule(formula=[formula2], fill=red_fill, font=red_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #--CON CONDICIONANTE: SI POTENCIA A FACTURAR FUE RECONOCIDO POTENCIA LEIDA BLOQUE ALTO TAMBIEN DEBE SER RECONOCIDO Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($AO2)), ISBLANK($AL2))', "AL2:AL1048576"),
        ('=AND(NOT(ISBLANK($AL2)), ISBLANK($AO2))', "AO2:AO1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #--CON CONDICIONANTE: SI ENERGIA REACTIVA FUE RECONOCIDO FACTOR DE POTENCIA TAMBIEN DEBE SER RECONOCIDA Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($AR2)), ISBLANK($AQ2))', "AQ2:AQ1048576"),
        ('=AND(NOT(ISBLANK($AQ2)), ISBLANK($AR2))', "AR2:AR1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #--CON CONDICIONANTE: SI EXCESO DE POTENCIA FUE RECONOCIDO EL IMPORTE POR EXCESO DE POTENCIA DEBER SERLO TAMBIEN Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($AY2)), ISBLANK($AP2))', "AP2:AP1048576"),
        ('=AND(NOT(ISBLANK($AP2)), ISBLANK($AY2))', "AY2:AY1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)
    
    #--CON CONDICIONANTE: SI POTENCIA LEIDA FUE RECONOCIDO MULTIPLICADOR DEBERIA SER RECONOCIDO Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($AL2)), ISBLANK($AC2))', "AC2:AC1048576"),
        ('=AND(NOT(ISBLANK($AL2)), ISBLANK($AL2))', "AL2:AL1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #############################################################################
    #-ESTAS REGLAS SON PARA LOS DATOS O CAMPOS QUE DEBEN SER EN FORMATO NUMERICO 
    formulas_rangos = [
        ('=AND(NOT(ISNUMBER($D2)), NOT(ISBLANK($D2)))', "D2:D1048576"),
        ('=AND(NOT(ISNUMBER($G2)), NOT(ISBLANK($G2)))', "G2:G1048576"),
        ('=AND(NOT(ISNUMBER($K2)), NOT(ISBLANK($K2)))', "K2:K1048576"),
        ('=AND(NOT(ISNUMBER(Q2)), NOT(ISBLANK(Q2)))', "Q2:AA1048576"),
        ('=AND(NOT(ISNUMBER(AC2)), NOT(ISBLANK(AC2)))', "AC2:BU1048576"),
        ('=AND(NOT(ISNUMBER($BW2)), NOT(ISBLANK($BW2)))', "BW2:BW1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=blue_fill, font=blue_font, border=blue_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=blue_fill, font=blue_font, border=blue_border)
        hoja.conditional_formatting.add(rango, rule)

    #############################################################################
    #-ESTAS REGLAS SON PARA LOS DATOS O CAMPOS QUE NO DEBEN SER NEGATIVOS 
    formulas_rangos = [
        ('=$D2<0', "D2:D1048576"),
        ('=$G2<0', "G2:G1048576"),
        ('=$K2<0', "K2:K1048576"),
        ('=Q2<0', "Q2:AA1048576"),
        ('=AC2<0', "AC2:AD1048576"),
        ('=AI2<0', "AI2:AR1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=lightgray_fill, font=blue_font, border=blue_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=lightgray_fill, font=blue_font, border=blue_border)
        hoja.conditional_formatting.add(rango, rule)

    #############################################################################
    #-ESTAS REGLAS SON PARA LOS DATOS O CAMPOS QUE NO DEBEN TENER DECIMALES
    formulas_rangos = [
        ('=ROUND($D2,0)<>$D2', "D2:D1048576"),
        ('=ROUND($G2,0)<>$G2', "G2:G1048576"),
        ('=ROUND($K2,0)<>$K2', "K2:K1048576"),
        ('=ROUND(Q2,0)<>Q2', "Q2:AA1048576"),
        ('=ROUND(AC2,0)<>AC2', "AC2:AQ1048576"),
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=blue_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=blue_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE Total energia a facturar = Energia consumida en n dias + Energia estimada + Energia adicional por cambio de medidor + Menos devolucion kWh + Perdidas en el transformador
    formula2 = '=$AI2<>($AD2+$AE2+$AF2+$AG2+$AH2)'
    rango2 = "AI2:AI1048576"
    rule2 = FormulaRule(formula=[formula2], font=blue_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE Importe por consumo = Importe por Cargo Fijo + Importe por Energia + Importe por Energia Bloque Alto + Importe por Energia Bloque Medio + Importe por Energia Bloque Bajo + Importe por potencia + Importe por exceso de potencia fuera de punta 
    formula2 = '=$AZ2<>($AS2+$AT2+$AU2+$AW2+$AV2+$AY2+$AX2)'
    rango2 = "AS2:AZ1048576"
    rule2 = FormulaRule(formula=[formula2], fill=yellow_fill, font=yellow_font, border=yellow_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE Importe total por el suministro = Importe total por el consumo + Mas debito consumo no facturado – Menos crédito por Devolucion + Importe por bajo factor de potencia + Mas cargo por conexión + Mas intereses por mora 
    formula2 = '=$BG2<>($BA2+$BE2-$BF2+$BB2+$BD2+$BC2)'
    rango2 = "BA2:BG1048576"
    rule2 = FormulaRule(formula=[formula2], fill=purple_fill, font=purple_font, border=purple_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE Importe  total factura = Importe total por el suministro + Tasas por alumbrado publico + Tasas por aseo urbano 
    formula2 = '=$BJ2<>($BG2+$BH2+$BI2)'
    rango2 = "BG2:BJ1048576"
    rule2 = FormulaRule(formula=[formula2], fill=brown_fill, font=brown_font, border=brown_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE Importe del mes a cancelar = Importe total factura + Mas crédito aplicado en cuenta corriente – Menos crédito anterior – Menos reducción calidad servicio técnico – Pagos adelantados aplicados + Mas depósito de garantía.
    formula2 = '=$BP2<>($BJ2+$BL2-$BM2-$BN2-$BO2+$BK2)'
    rango2 = "BJ2:BP1048576"
    rule2 = FormulaRule(formula=[formula2], fill=orange_fill, font=orange_font, border=orange_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE Importe total a cancelar = Importe del mes a cancelar + Mas deudas pendientes de energía + Deudas pendientes de tasa de aseo – Menos pago adelantado
    formula2 = '=$BU2<>($BP2+$BQ2+$BR2-$BS2)'
    rango2 = "BP2:BU1048576"
    rule2 = FormulaRule(formula=[formula2], fill=green_fill, font=green_font, border=green_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE EL DATO DE CONSUMIDOR ES CORRECTO
    formula2 = '=$F2<>$E2'
    rango2 = "E2:F1048576"
    rule2 = FormulaRule(formula=[formula2], font=red_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE EL DATO DE TOTAL ENERGIA A FACTURAR
    formula2 = '=$AI2<>$AJ2'
    rango2 = "AI2:AJ1048576"
    rule2 = FormulaRule(formula=[formula2], font=red_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #-ESTA REGLA ES PARA VERIFICAR QUE EL DATO DE CONSUMIDOR ES CORRECTO
    formulas_rangos = [
        ('=$BW2<>$BU2', "BU2:BU1048576"),
        ('=$BW2<>$BU2', "BW2:BW1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    # Guardar el archivo de Excel con los cambios
    libro.save(file_path)
########################################## FIN DE LA SECCION 1 ##########################################

########################################## SECCION 2: CREACION O APERTURA DE CARPETAS Y MODELOS DE IA ##########################################
ruta_destino = 'img_of_pdf'             #Se extrae las facturas del pdf en la carpeta "img_of_pdf" 
if not os.path.exists(ruta_destino):    #en el caso de que la carpeta no exista 
    os.makedirs(ruta_destino)           #entonces se crea la carpeta

model = YOLO("best_ultimo.pt")    # Carga el modelo entrenado

#Cargamos las librerias para el reconocimiento OCR
reader_es = easyocr.Reader(['es'], gpu=True)  
device = torch.device("mps" if torch.backends.mps.is_available() else "cuda" if torch.cuda.is_available() else "cpu")
model_version="microsoft/trocr-small-printed"
processor=TrOCRProcessor.from_pretrained(model_version)
modelocr=VisionEncoderDecoderModel.from_pretrained(model_version).to(device)
########################################## FIN DE LA SECCION 2 ##########################################

########################################## FUNCION 1: PROGRAMA DE EXTRACCION DE DATOS USANDO IA ##########################################
def sistema_de_lectura():
    
    pdf_files = glob.glob("PDFs_file/*.PDF")    # Obtener la lista de archivos con extensión .PDF

    for pdf_file in pdf_files:
        pdf_document = fitz.open(pdf_file)      #Preprocesamiento del
        zoom_x = 3.0                            #.PDF para que las imagenes
        zoom_y = 3.0                            #sean extraidas
        mat = fitz.Matrix(zoom_x, zoom_y)       #con mejor calidad
    
        for i, page in enumerate(pdf_document, start=1):                        #Cada pagina del PDF
            pix = page.get_pixmap(matrix=mat, alpha=False)                      #la guardamos como imagen
            nombre_archivo = os.path.join(ruta_destino, f"factura{i}.jpg")      #con el nombre de su respectivo orden
            pix.save(nombre_archivo)                                            #guardando las imágenes en "img_of_pdf"
        
        pdf_document.close()  # Cerramos el archivo PDF
        
        datos_totales = []      #Los datos extraidos se guardan en este vector
        
        def extract_number(filename):                               #Funcion para ordenar numericamente.
            match = re.search(r'\d+', filename)                     #Retorna inf si no hay número 
            return int(match.group()) if match else float('inf')    #para que estos archivos vayan al final

        image_paths = sorted(glob.glob(os.path.join(ruta_destino, "*.jpg")), key=lambda x: extract_number(os.path.basename(x)))     #Lista de todas las imágenes ordenadas numéricamente

        #Itera sobre todas las imágenes en el directorio que contiene las imágenes
        for image_path in image_paths:
            
            results = model.predict(source=image_path, conf=0.5, save=False, line_width=2, show_labels=False, device=device, imgsz=864)      # Detecta los campos de la imagen

            # Iterar sobre los resultados de la detección
            for result in results:
                boxes = result.boxes.cpu().numpy()  # Obtenemos las coordenadas de los campos reconocidos
                coords = boxes.xywhn                # coordenadas en el formato xywhn
                clases = boxes.cls                  # y nombres de los campos reconocidos

            image_easyocr = cv2.imread(image_path)
            height, width, _ = image_easyocr.shape

            textos_por_clase = [""] * 75

            # Iterar sobre un respectivo campo reconocido
            for coord, clase in zip(coords, clases):
                
                # Convertir coordenadas normalizadas a coordenadas absolutas
                x_center = float(coord[0]) * width
                y_center = float(coord[1]) * height
                w = float(coord[2]) * width
                h = float(coord[3]) * height

                # Calcular las coordenadas de las esquinas del rectángulo
                x = int(x_center - w / 2)
                y = int(y_center - h / 2)
                x2 = int(x_center + w / 2)
                y2 = int(y_center + h / 2)

                #Algunos errores recurrentes que se deben filtrar, eliminar o corregir
                corrections = {
                r" =|~|\*|:| =|: |\. ": "",
                r'"': '',
                "=": "",
                "CONSTRUCGION": "CONSTRUCCION",
                "EETG |EETG.|EETG|EETC|EETC |E. ET.C. ": "EETC ",
                "GI-": "G1-",
                "LAPAZ": "LA PAZ",
                "ELALTO": "EL ALTO",
                "-GID-|-GDD-|-GP-":"-GD-",
                "-ABP-|-AEF-|-AER-":"-ABR-",
                "-OCI-|-OCL-":"-OCT-",
                "-AG0-":"-AGO-",
                "TELEFEERICO|TELEFEFRICO|TELEFERRICO|TELEFEERICO":"TELEFERICO",
                "ABFIL|ABRL|ABRII|ABFL|ABELL|AERIL":"ABRIL",
                "JUNIC":"JUNIO",
                "JULIC|JULI0":"JULIO",
                "SEPTIERBRE|SEPTIMEMBRE|SEPTIERMBRE|SEPTIEMBERE|SEPTEMBERE|SEPTEMBRE|SEPTIMBRE|SEPTLEMBRE":"SEPTIEMBRE",
                "DICIERBRE|DICLEMBRE|DICIERMBRE|DICIMBRE|DICIERNBRE|DICIERHBRE|DICEMBRE":"DICIEMBRE",
                "NOVIERBRE|NOVIMBRE|NOVIEWBRE":"NOVIEMBRE",
                "ESTAGION":"ESTACION",
                "UCELESTE":"L/CELESTE",
                "UAMARILLA,LAMARAILLA,UJAMARILLA,LJAMARILLA":"L/AMARILLA",
                "2800048029":"280048029",
                "985888-1-5":"985688-1-5",
                "985888-5":"985688-5",
                "LECTURA NORMAT|LECTURA NORMAI|LECTURA NOMMAL|LECTURA NOMAL|LECTURANORMAL":"LECTURA NORMAL",
                "LECLURA":"LECTURA",
                r"\bO KWH\b|\bO KWN\b|\bO KVARH\b|\bO KVARN\b|\bO KVAFN\b|\bO KW\b": "0",
                r"\bOKWH\b|\bOKWN\b|\bOKVARH\b|\bOKVARN\b|\bOKVAFN\b|\bOKW\b": "0",
                " KWH| KWN| KVARH| KVARN| KVAFN| KW|POTENCIA |PATENCIA |POLENCIA |POTENCLA ":"",
                "KWH|KWN|KVARH|KVARN|KVAFN|KW|POTENCIA|PATENCIA|POLENCIA|POTENCLA":"",
                "ENTE ":"ENFE "
                }

                #Si el texto es de "Actividad" o "Direccion" usamos EasyOCR
                if clase==0 or clase==15:
                    
                    roi_image = image_easyocr[y:y2, x:x2]   # Recortar la región de la imagen
                    
                    (height1, width1, canal) = roi_image.shape                          #Redimensionamos
                    redim_roi_image = cv2.resize(roi_image, (width1 * 7, height1 * 7))  #la imagen

                    denoised_image = cv2.GaussianBlur(redim_roi_image, (5, 5), 0)   #Aplicamos reducción de ruido

                    kernel_sharpening = np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]]) #Aumentamos
                    sharpened_image = cv2.filter2D(denoised_image, -1, kernel_sharpening)   #la nitidez

                    gray_image = cv2.cvtColor(sharpened_image, cv2.COLOR_BGR2GRAY)              #Imagen en blanco y negro
                    _, binary_image = cv2.threshold(gray_image, 150, 255, cv2.THRESH_TOZERO)    #Binarización

                    valor = reader_es.readtext(binary_image, detail=0, paragraph=True)  #Extraemos el texto

                    for text in valor:                                          #Si se detecta uno de las
                        for pattern, replacement in corrections.items():        #errores mostrados anteriormente
                            text = re.sub(pattern, replacement, text)           #se lo corrige
                    valor=text                                                  #y se lo exporta a Excel

                #Si el texto es de cualquier otra clase usamos TrOCR    
                else:
                    image = Image.open(image_path).convert("RGB")   #Preprocesa la imagen
                    crp_image=image.crop((x,y,x2,y2))               #y la seccion que nos interesa
                    
                    pixel_values=processor(crp_image, return_tensors="pt").pixel_values.to(device)      #Corta la imagen
                    generated_ids = modelocr.generate(pixel_values, max_new_tokens=50)                  #en la seccion que nos interesa
                    text_easy_ocr=processor.batch_decode(generated_ids, skip_special_tokens=True)[0]    #y extrae el texto

                    for pattern, replacement in corrections.items():                    #Si hay error 
                        text_easy_ocr = re.sub(pattern, replacement, text_easy_ocr)     #lo corrige

                    #Se asigna el respectivo tipo de valor a cada texto extraido (Texto o numero)
                    texto_sin_puntos = text_easy_ocr.replace(",", "").replace(".", "").replace(" - ","").replace(" -","").replace("- ","").replace("-","")
                    if texto_sin_puntos.isnumeric():
                        if ',' in text_easy_ocr or '.' in text_easy_ocr:
                            if text_easy_ocr.count('.') > 1:                #Si hay más de un punto
                                partes = text_easy_ocr.split('.')           #elimina el primero
                                texto_sin_comas = ''.join(partes[:-1]) + '.' + partes[-1]
                            elif text_easy_ocr.count('.') == 0 and text_easy_ocr.count(',') == 1:
                                texto_sin_comas = text_easy_ocr.replace(",", ".") 
                            else:
                                texto_sin_comas = text_easy_ocr.replace(",", "")  #Elimina todas las comas
                            try:
                                valor = float(texto_sin_comas)
                            except ValueError:
                                valor = text_easy_ocr
                        elif '-' in text_easy_ocr:
                            valor=text_easy_ocr.replace(" - ","-").replace(" -","-").replace("- ","-").replace("-","-")
                        else:
                            valor = int(text_easy_ocr)
                    else:
                        valor = text_easy_ocr

                mapeo_clase_a_posicion = [9, 19, 18, 20, 25, 24, 26, 22, 21, 23, 10, 13, 8, 4, 69, 7, 30, 42, 31, 29, 41, 43, 1, 15, 14, 74, 67, 44, 51, 45, 46, 48, 47, 50, 49, 53, 72, 61, 52, 58, 17, 16, 55, 56, 62, 68, 54, 6, 57, 32, 70, 65, 12, 28, 3, 71, 66, 33, 36, 37, 39, 38, 40, 11, 73, 2, 59, 60, 27, 34, 35, 63, 64, 5]    #Lista de mapeo que define cómo los valores de clase corresponden a las posiciones en textos_por_clase

                textos_por_clase[mapeo_clase_a_posicion[int(clase)]] = valor    #Asigna el valor a la posición correspondiente en textos_por_clase usando el mapeo
                
            textos_por_clase[0] = re.sub('img_of_pdf/|.jpg', '',re.sub('PDFs_file/','',pdf_file) + "_" + image_path )   #La primera celda de la fila es el nombre del PDF + el numero de imagen que fue procesada
            datos_totales.append(textos_por_clase)  #Se van guardando todos los textos y numeros extraidos

        fila_inicio = hoja.max_row + 1  #Encontrar la primera fila vacía

        #Escribir los datos en la hoja
        for fila, datos_imagen in enumerate(datos_totales, start=fila_inicio):
            for columna, dato_clase in enumerate(datos_imagen, start=1):
                if isinstance(dato_clase, list):
                    valor = ', '.join(map(str, dato_clase))  #Convertir la lista a una cadena separada por comas
                else:
                    valor = dato_clase
                celda = hoja.cell(row=fila, column=columna)
                celda.value = valor

        for image_path in glob.glob(os.path.join(ruta_destino, "*.jpg")):   #Elimina las imágenes procesadas
            os.remove(image_path)                                           #de img_of_pdf

    hoja.freeze_panes = hoja['B2']  #Inmoviliza paneles
    libro.save(file_path)           #Guardamos el libro de trabajo
    libro.close()                   #Cerramos el libro
    
    for pdf_file in pdf_files:      #Eliminamos los PDF procesados
        os.remove(pdf_file)

    #
    # Cargar el libro de trabajo y la hoja de datos
    wb = load_workbook(file_path)
    ws = wb.active  # Asume que quieres trabajar con la hoja activa

    # Verificar si ya existe una tabla en la hoja
    table_exists = any(
        isinstance(table, Table) for table in ws._tables.values()
    )

    if not table_exists:
        # Crear una tabla en la hoja de datos (asegúrate de que los datos están en un rango apropiado)
        tab = Table(displayName='TablaDatos', ref=ws.dimensions)

        # Agregar estilo a la tabla con colores personalizados
        style = TableStyleInfo(
            name='TableStyleMedium9',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )

        # Aplicar color a las filas y columnas de la tabla
        # Usar el formato aRGB para los colores
        header_fill = PatternFill(start_color="FFB1A0C7", end_color="FFB1A0C7", fill_type="solid")  # Morado
        even_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # Blanco
        odd_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # Blanco

         # Definir el borde plomo
        border_color = "D0D0D0"  # Gris claro
        border_side = Side(border_style="thin", color=border_color)
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

         # Aplicar color de fondo y borde a las celdas de la tabla
        # Aplicar color a las celdas de la tabla
        for row in ws[tab.ref]:
            for cell in row:
                if cell.row == ws[tab.ref][0][0].row:  # Si es la fila de encabezado
                    cell.fill = header_fill
                elif cell.row % 2 == 0:  # Filas pares
                    cell.fill = even_fill
                else:  # Filas impares
                    cell.fill = odd_fill
                cell.border = border

        # Agregar el estilo a la tabla
        tab.tableStyleInfo = style

        # Añadir la tabla al worksheet
        ws.add_table(tab)

    # Guardar el archivo
    wb.save(file_path)
    #

    # Actualizar el mensaje con la ruta relativa
    current_dir = os.getcwd()                                                                       #Obtenemos la ruta actual
    relative_path = os.path.relpath(file_path, current_dir)                                         #la convertimos en ruta relativa
    mensaje_datos_extraidos = "Los datos de las facturas han sido extraidos correctamente en:"      #para mostrar donde se 
    canvas.itemconfig(mensaje_datos_extraidos_id, text=mensaje_datos_extraidos)                     #guardo el Excel
    mensaje_datos_extraidos2 = f"{relative_path}"                                                   #con los datos extraidos 
    canvas.itemconfig(mensaje_datos_extraidos2_id, text=mensaje_datos_extraidos2)                   #e indicar el fin de la extraccion
    pass
########################################## FIN DE FUNCION 1 ##########################################

########################################## FUNCION 2: PROGRAMA PARA MOSTRAR MENSAJES DE INICIO Y FIN EN LA VENTANA DEL PROGRAMA ##########################################
def iniciar_lectura():
    mensaje = "Espere mientras se extraen los datos..."     #Este mensaje se muestra cuando se
    canvas.itemconfig(mensaje_id, text=mensaje)             #inicia el proceso de extraer los datos

    mensaje_datos_extraidos = ""                                                #Cuando se estan extrayendo nuevos datos
    canvas.itemconfig(mensaje_datos_extraidos_id, text=mensaje_datos_extraidos) #se limpia este mensaje

    mensaje_datos_extraidos2 = ""                                                   #Cuando se estan extrayendo nuevos datos
    canvas.itemconfig(mensaje_datos_extraidos2_id, text=mensaje_datos_extraidos2)   #se limpia este mensaje

    ventana.update()    #Actualiza la interfaz gráfica para mostrar el mensaje

    sistema_de_lectura()    #Iniciamos el proceso de extraccion de datos
    
    mensaje = ""                                    #Cuando se terminan de extraer los datos
    canvas.itemconfig(mensaje_id, text=mensaje)     #se limpia el mensaje de "Espere mientras se extraen los datos..."
########################################## FIN DE FUNCION 2 ##########################################

########################################## FUNCION 3: PROGRAMA PARA CARGAR ARCHIVO ##########################################
def cargar_archivo():
    archivo = filedialog.askopenfilename()
    if archivo:
        guardar_en_carpeta(archivo)
########################################## FIN DE FUNCION 3 ##########################################

########################################## FUNCION 4: PROGRAMA PARA GUARDAR EL ARCHIVO CARGADO ##########################################
def guardar_en_carpeta(ruta_archivo):
    carpeta_destino = "PDFs_file"               
    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)
    
    nombre_archivo = os.path.basename(ruta_archivo)
    ruta_destino = os.path.join(carpeta_destino, nombre_archivo)
    
    try:
        shutil.copy(ruta_archivo, ruta_destino)
        print(f"Archivo copiado en: {ruta_destino}")
        mensaje_cargar = f"{nombre_archivo} cargado correctamente"
        canvas.itemconfig(mensaje_cargar_id, text=mensaje_cargar)

    except Exception as e:
        print(f"Error al copiar el archivo: {e}")
        mensaje_cargar = f"Error al cargar el archivo: {e}"
        canvas.itemconfig(mensaje_cargar_id, text=mensaje_cargar)
########################################## FIN DE FUNCION 4 ##########################################

########################################## FUNCION 5: PROGRAMA PARA ELIMINAR ARCHIVOS CUANDO SE CIERRA LA VENTANA ##########################################
def limpiar_carpetas():
    carpetas = ["img_of_pdf", "PDFs_file"]
    for carpeta in carpetas:
        if os.path.exists(carpeta):
            for archivo in os.listdir(carpeta):
                archivo_path = os.path.join(carpeta, archivo)
                try:
                    if os.path.isfile(archivo_path) or os.path.islink(archivo_path):
                        os.unlink(archivo_path)
                    elif os.path.isdir(archivo_path):
                        shutil.rmtree(archivo_path)
                except Exception as e:
                    print(f'Error al borrar {archivo_path}. Razón: {e}')

def on_closing():
    limpiar_carpetas()
    ventana.destroy()   
########################################## FIN DE FUNCION 5 ##########################################

####################INICIO DE LA INTERFAZ DE LA VENTANA PRINCIPAL##################################
ventana = tk.Tk()                                       #Creamos la
ventana.title("LECTOR DE FACTURAS ELECTRICAS")          #ventana principal

ancho_pantalla = ventana.winfo_screenwidth()            #Obtenemos las dimensiones de 
alto_pantalla = ventana.winfo_screenheight()            #la pantalla del dispositivo
ancho_ventana = 800                                     #para tener un ancho
alto_ventana = 600                                      #y alto de la ventana
posicion_x = (ancho_pantalla - ancho_ventana) // 2      #centrada en medio
posicion_y = (alto_pantalla - alto_ventana) // 2        #de la pantalla

ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{posicion_x}+{posicion_y}")   #Establecemos el tamaño y la posición de la ventana

canvas = tk.Canvas(ventana, width=ancho_ventana, height=alto_ventana)           #Creamos un canvas
canvas.pack(fill="both", expand=True)                                           #para una interfaz mas amigable

imagen_fondo = Image.open("background.png")                                                                                     #Estableciendo
imagen_fondo = imagen_fondo.resize((ancho_ventana, alto_ventana), Image.Resampling.LANCZOS)                                     #un background
imagen_fondo = ImageTk.PhotoImage(imagen_fondo)                                                                                 #de la
canvas.create_image(0, 0, anchor="nw", image=imagen_fondo)                                                                      #EETC MT

#Texto y botones que se muestran en la ventana
fuente_texto=("times", 25, "bold")

canvas.create_text(ancho_ventana//2, 50, text="Bienvenido al lector automatizado de facturas del servicio electrico", font=fuente_texto, fill="black")  #Cabecera

btn_seleccionar = tk.Button(ventana, text="Cargue su PDF", font=fuente_texto, command=cargar_archivo)   #Boton 1
canvas.create_window(ancho_ventana//2, 150, window=btn_seleccionar)

mensaje_cargar = ""                                                                                                     #Texto 1
mensaje_cargar_id = canvas.create_text(ancho_ventana // 2, 200, text=mensaje_cargar, font=fuente_texto, fill="black")

mensaje = ""                                                                                            #Texto 2
mensaje_id = canvas.create_text(ancho_ventana//2, 250, text=mensaje, font=fuente_texto, fill="black")

mensaje_datos_extraidos = ""                                                                                                            #Texto 3
mensaje_datos_extraidos_id = canvas.create_text(ancho_ventana//2, 300, text=mensaje_datos_extraidos, font=fuente_texto, fill="black")

mensaje_datos_extraidos2 = ""                                                                                                           #Texto 4
mensaje_datos_extraidos2_id = canvas.create_text(ancho_ventana//2, 350, text=mensaje_datos_extraidos2, font=fuente_texto, fill="black")

btn_iniciar = tk.Button(ventana, text="Inicie el programa", font=fuente_texto, command=iniciar_lectura) #Boton 2
canvas.create_window(ancho_ventana//2, 400, window=btn_iniciar)

canvas.create_text(ancho_ventana//2, 550, text="Desarrollado por JaRs", font=fuente_texto, fill="black")    #Copyright

ventana.protocol("WM_DELETE_WINDOW", on_closing)    #Protocolo al cerrar la ventana

ventana.mainloop()      # Ejecutar el bucle principal de tkinter
###########################FIN de la interfaz##########################################
