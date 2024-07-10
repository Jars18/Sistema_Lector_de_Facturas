
##ESTE PROGRAMA SOLO ES UN EJECUTABLE DE PYTHON
import os
from pdf2image import convert_from_path
from ultralytics import YOLO
import glob
import re
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
from PIL import Image
import easyocr
import cv2
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule
import torch
import sys
##########################################INICIO DEL SISTEMA DE LECTURA##########################################
# Crear o abrir el libro de trabajo

# Obtener el directorio del ejecutable
if getattr(sys, 'frozen', False):
    # PyInstaller añade este atributo
    application_path = os.path.dirname(sys.executable)
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

# Definir la ruta completa del archivo de Excel
file_path = os.path.join(application_path, "Informacion.xlsx")
                         
if os.path.exists(file_path):
    libro = openpyxl.load_workbook(file_path)
    hoja = libro.active
else:
    libro = Workbook()
    hoja = libro.active

    # Definir encabezados
    encabezados = [
        "factura", "Fecha", "Senor", "NIT-CI", "Consumidor", "val_consumidor", "Medidor",
        "Direccion", "Ciudad o Localidad", "Actividad", "Carta Factura", "Remesa y Ruta", "Mes",
        "Categoria Tarifaria", "Fecha de Lectura Anterior", "Fecha de Lectura Actual",
        "Lectura Medidor Anterior", "Lectura Medidor Actual", "Bloque Alto Anterior",
        "Bloque Alto Actual", "Bloque Alto kWh", "Bloque Medio Anterior", "Bloque Medio Actual",
        "Bloque Medio kWh", "Bloque Bajo Anterior", "Bloque Bajo Actual", "Bloque Bajo kWh",
        "Tipo de Lectura", "Multiplicador", "Energia consumida en -n- dias", "Energia Estimada",
        "Energia adicional por cambio de medidor", "Menos devolucion kWh", "Perdidas en el transformador",
        "Total energia a facturar", "Total energia a facturar valid", "Potencia Contratada",
        "Potencia Leida Bloque Alto", "Potencia Leida Bloque Medio", "Potencia Leida Bloque Bajo",
        "Potencia a Facturar", "Exceso de potencia fuera de punta", "Energia Reactiva", "Factor de potencia",
        "Importe por Cargo Fijo", "Importe por Energia", "Importe por Energia - Bloque Alto",
        "Importe por Energia - Bloque Medio", "Importe por Energia - Bloque Bajo", "Importe por Potencia",
        "Importe por Exceso de potencia fuera de punta", "Importe por Consumo", "Importe total por consumo",
        "Importe por bajo factor de potencia", "Mas intereses por Mora", "Mas cargo por conexion",
        "Mas debito consumo no facturado", "Menos credito por devolucion", "Importe total por el suministro",
        "Tasas por Alumbrado Publico", "Tasas por Aseo Urbano", "Importe total factura", "Mas deposito de garantia",
        "mas credito aplicado en cuenta corriente", "menos credito anterior", "Menos reduccion calidad servicio tecnico",
        "Pagos adelantados aplicados", "Importe del mes a cancelar", "Mas deudas pendientes de energia",
        "Deudas pendientes de tasa de aseo", "Menos pago adelantado", "Pago adelantado a la fecha",
        "Importe total a cancelar", "Resumen mes y ano", "Importe"
    ]

    # Insertar encabezados en la primera fila
    for col_num, header in enumerate(encabezados, 1):
        hoja.cell(row=1, column=col_num, value=header)

    # Definir el formato condicional
    red_fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
    red_font = Font(color="9c0006")
    red_border = Border(
        left=Side(style='thin', color='9c0006'),
        right=Side(style='thin', color='9c0006'),
        top=Side(style='thin', color='9c0006'),
        bottom=Side(style='thin', color='9c0006')
    )


    blue_fill = PatternFill(start_color="b8cce4", end_color="b8cce4", fill_type="solid")
    blue_font = Font(color="1f497d")
    blue_border = Border(
        left=Side(style='thin', color='1f497d'),
        right=Side(style='thin', color='1f497d'),
        top=Side(style='thin', color='1f497d'),
        bottom=Side(style='thin', color='1f497d')
    )

    lightgray_fill = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")

    yellow_fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
    yellow_font = Font(color="9c5700")
    yellow_border = Border(
        left=Side(style='thin', color='9c5700'),
        right=Side(style='thin', color='9c5700'),
        top=Side(style='thin', color='9c5700'),
        bottom=Side(style='thin', color='9c5700')
    )

    purple_fill = PatternFill(start_color="ccc0da", end_color="ccc0da", fill_type="solid")
    purple_font = Font(color="60497a")
    purple_border = Border(
        left=Side(style='thin', color='60497a'),
        right=Side(style='thin', color='60497a'),
        top=Side(style='thin', color='60497a'),
        bottom=Side(style='thin', color='60497a')
    )

    brown_fill = PatternFill(start_color="c4bd97", end_color="c4bd97", fill_type="solid")
    brown_font = Font(color="ab7942")
    brown_border = Border(
        left=Side(style='thin', color='ab7942'),
        right=Side(style='thin', color='ab7942'),
        top=Side(style='thin', color='ab7942'),
        bottom=Side(style='thin', color='ab7942')
    )

    orange_fill = PatternFill(start_color="fcd5b4", end_color="fcd5b4", fill_type="solid")
    orange_font = Font(color="e26b0a")
    orange_border = Border(
        left=Side(style='thin', color='e26b0a'),
        right=Side(style='thin', color='e26b0a'),
        top=Side(style='thin', color='e26b0a'),
        bottom=Side(style='thin', color='e26b0a')
    )

    green_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
    green_font = Font(color="006100")
    green_border = Border(
        left=Side(style='thin', color='006100'),
        right=Side(style='thin', color='006100'),
        top=Side(style='thin', color='006100'),
        bottom=Side(style='thin', color='006100')
    )

    # Definir las fórmulas y rangos
    #ESTAS REGLAS SON PARA LOS DATOS O CAMPOS QUE NO FUERON RECONOCIDOS Y DEBIERON SERLO
    #DE FORMA GENERAL
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(B2))', "B2:I1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(L2))', "L2:P1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(AB2))', "AB2:AB1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(AD2))', "AD2:AD1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(AI2))', "AI2:AJ1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(AZ2))', "AZ2:BA1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(BG2))', "BG2:BG1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(BJ2))', "BJ2:BJ1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(BP2))', "BP2:BP1048576"),
        ('=AND(NOT(ISBLANK($A2)), ISBLANK(BU2))', "BU2:BW1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #CON CONDICIONANTE: SI LECTURA MEDIDOR ANTERIOR FUE RECONOCIDO TAMBIEN DEBE SER RECONOCIDO LECTURA MEDIDOR ACTUAL Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($Q2)), ISBLANK($R2))', "R2:R1048576"),
        ('=AND(NOT(ISBLANK($R2)), ISBLANK($Q2))', "Q2:Q1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #CON CONDICIONANTE: SI UNA LECTURA DE BLOQUE FUE IDENTIFICADO TODOS LOS DEMAS CAMPOS DE LECTURA DE BLOQUE TAMBIEN DEBERIAN SER IDENTIFICADOS
    formula2 = '=AND(SUM(--(NOT(ISBLANK($T2:$W2)))) > 0, ISBLANK(S2))'
    rango2 = "S2:AA1048576"
    rule2 = FormulaRule(formula=[formula2], fill=red_fill, font=red_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #CON CONDICIONANTE: SI UNA POTENCIA LEIDA DE BLOQUE BAJO O MEDIO FUE IDENTIFICADA TODOS LOS CAMPOS DE POTENCIA LEIDA DEBEN SER RECONOCIDO
    formula2 = '=AND(OR(NOT(ISBLANK($AN2)), NOT(ISBLANK($AM2))), OR(ISBLANK($AM2), ISBLANK($AL2), ISBLANK($AN2)))'
    rango2 = "AL2:AN1048576"
    rule2 = FormulaRule(formula=[formula2], fill=red_fill, font=red_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #CON CONDICIONANTE: SI POTENCIA A FACTURAR FUE RECONOCIDO POTENCIA LEIDA BLOQUE ALTO TAMBIEN DEBE SER RECONOCIDO Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($AO2)), ISBLANK($AL2))', "AL2:AL1048576"),
        ('=AND(NOT(ISBLANK($AL2)), ISBLANK($AO2))', "AO2:AO1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #CON CONDICIONANTE: SI ENERGIA REACTIVA FUE RECONOCIDO FACTOR DE POTENCIA TAMBIEN DEBE SER RECONOCIDA Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($AR2)), ISBLANK($AQ2))', "AQ2:AQ1048576"),
        ('=AND(NOT(ISBLANK($AQ2)), ISBLANK($AR2))', "AR2:AR1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #CON CONDICIONANTE: SI EXCESO DE POTENCIA FUE RECONOCIDO EL IMPORTE POR EXCESO DE POTENCIA DEBER SERLO TAMBIEN Y VICEVERSA
    formulas_rangos = [
        ('=AND(NOT(ISBLANK($AY2)), ISBLANK($AP2))', "AP2:AP1048576"),
        ('=AND(NOT(ISBLANK($AP2)), ISBLANK($AY2))', "AY2:AY1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=red_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=red_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #############################################################################
    #ESTAS REGLAS SON PARA LOS DATOS O CAMPOS QUE DEBEN SER EN FORMATO NUMERICO 
    formulas_rangos = [
        ('=AND(NOT(ISNUMBER($D2)), NOT(ISBLANK($D2)))', "D2:D1048576"),
        ('=AND(NOT(ISNUMBER($G2)), NOT(ISBLANK($G2)))', "G2:G1048576"),
        ('=AND(NOT(ISNUMBER($K2)), NOT(ISBLANK($K2)))', "K2:K1048576"),
        ('=AND(NOT(ISNUMBER(Q2)), NOT(ISBLANK(Q2)))', "Q2:AA1048576"),
        ('=AND(NOT(ISNUMBER(AC2)), NOT(ISBLANK(AC2)))', "AC2:BU1048576"),
        ('=AND(NOT(ISNUMBER($BW2)), NOT(ISBLANK($BW2)))', "BW2:BW1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=blue_fill, font=blue_font, border=blue_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=blue_fill, font=blue_font, border=blue_border)
        hoja.conditional_formatting.add(rango, rule)

    #############################################################################
    #ESTAS REGLAS SON PARA LOS DATOS O CAMPOS QUE NO DEBEN SER NEGATIVOS 
    formulas_rangos = [
        ('=$D2<0', "D2:D1048576"),
        ('=$G2<0', "G2:G1048576"),
        ('=$K2<0', "K2:K1048576"),
        ('=Q2<0', "Q2:AA1048576"),
        ('=AC2<0', "AC2:AD1048576"),
        ('=AI2<0', "AI2:AR1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=lightgray_fill, font=blue_font, border=blue_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=lightgray_fill, font=blue_font, border=blue_border)
        hoja.conditional_formatting.add(rango, rule)

    #############################################################################
    #ESTAS REGLAS SON PARA LOS DATOS O CAMPOS QUE NO DEBEN TENER DECIMALES
    formulas_rangos = [
        ('=ROUND($D2,0)<>$D2', "D2:D1048576"),
        ('=ROUND($G2,0)<>$G2', "G2:G1048576"),
        ('=ROUND($K2,0)<>$K2', "K2:K1048576"),
        ('=ROUND(Q2,0)<>Q2', "Q2:AA1048576"),
        ('=ROUND(AC2,0)<>AC2', "AC2:AQ1048576"),
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], fill=blue_fill, font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], fill=blue_fill, font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE Total energia a facturar = Energia consumida en n dias + Energia estimada + Energia adicional por cambio de medidor + Menos devolucion kWh + Perdidas en el transformador
    formula2 = '=$AI2<>($AD2+$AE2+$AF2+$AG2+$AH2)'
    rango2 = "AI2:AI1048576"
    rule2 = FormulaRule(formula=[formula2], font=blue_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE Importe por consumo = Importe por Cargo Fijo + Importe por Energia + Importe por Energia Bloque Alto + Importe por Energia Bloque Medio + Importe por Energia Bloque Bajo + Importe por potencia + Importe por exceso de potencia fuera de punta 
    formula2 = '=$AZ2<>($AS2+$AT2+$AU2+$AW2+$AV2+$AY2+$AX2)'
    rango2 = "AS2:AZ1048576"
    rule2 = FormulaRule(formula=[formula2], fill=yellow_fill, font=yellow_font, border=yellow_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE Importe total por el suministro = Importe total por el consumo + Mas debito consumo no facturado – Menos crédito por Devolucion + Importe por bajo factor de potencia + Mas cargo por conexión + Mas intereses por mora 
    formula2 = '=$BG2<>($BA2+$BE2-$BF2+$BB2+$BD2+$BC2)'
    rango2 = "BA2:BG1048576"
    rule2 = FormulaRule(formula=[formula2], fill=purple_fill, font=purple_font, border=purple_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE Importe  total factura = Importe total por el suministro + Tasas por alumbrado publico + Tasas por aseo urbano 
    formula2 = '=$BJ2<>($BG2+$BH2+$BI2)'
    rango2 = "BG2:BJ1048576"
    rule2 = FormulaRule(formula=[formula2], fill=brown_fill, font=brown_font, border=brown_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE Importe del mes a cancelar = Importe total factura + Mas crédito aplicado en cuenta corriente – Menos crédito anterior – Menos reducción calidad servicio técnico – Pagos adelantados aplicados + Mas depósito de garantía.
    formula2 = '=$BP2<>($BJ2+$BL2-$BM2-$BN2-$BO2+$BK2)'
    rango2 = "BJ2:BP1048576"
    rule2 = FormulaRule(formula=[formula2], fill=orange_fill, font=orange_font, border=orange_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE Importe total a cancelar = Importe del mes a cancelar + Mas deudas pendientes de energía + Deudas pendientes de tasa de aseo – Menos pago adelantado
    formula2 = '=$BU2<>($BP2+$BQ2+$BR2-$BS2)'
    rango2 = "BP2:BU1048576"
    rule2 = FormulaRule(formula=[formula2], fill=green_fill, font=green_font, border=green_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE EL DATO DE CONSUMIDOR ES CORRECTO
    formula2 = '=$F2<>$E2'
    rango2 = "E2:F1048576"
    rule2 = FormulaRule(formula=[formula2], font=red_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE EL DATO DE TOTAL ENERGIA A FACTURAR
    formula2 = '=$AI2<>$AJ2'
    rango2 = "AI2:AJ1048576"
    rule2 = FormulaRule(formula=[formula2], font=red_font, border=red_border)
    hoja.conditional_formatting.add(rango2, rule2)

    #############################################################################
    #ESTA REGLA ES PARA VERIFICAR QUE EL DATO DE CONSUMIDOR ES CORRECTO
    formulas_rangos = [
        ('=$BW2<>$BU2', "BU2:BU1048576"),
        ('=$BW2<>$BU2', "BW2:BW1048576")
    ]
    rule1 = FormulaRule(formula=[formulas_rangos[0][0]], font=red_font, border=red_border)
    for formula, rango in formulas_rangos:
        rule = FormulaRule(formula=[formula], font=red_font, border=red_border)
        hoja.conditional_formatting.add(rango, rule)

    # Guardar el archivo de Excel con los cambios
    libro.save(file_path)


ruta_destino = 'img_of_pdf'             #Se extrae las facturas del pdf en la carpeta "img_of_pdf" 
if not os.path.exists(ruta_destino):    #en el caso de que la carpeta no exista 
    os.makedirs(ruta_destino)           #entonces se crea la carpeta

#El modelo usado antes de 74clases es modelo_espacios_mas_especificos.pt, que dio buenos resultados
model = YOLO("/Users/jurgenalejandrorocasalvosanchez/Documents/Programa_PDG/Lector_de_Facturas/best_ultimo.pt")    # Carga el modelo entrenado
#model = YOLO("/Users/jurgenalejandrorocasalvosanchez/Documents/Programa_PDG/Lector_de_Facturas/best_specific_spaces_1000img-74clases.pt")    # Carga el modelo entrenado

#Cargamos las librerias para el reconocimiento OCR
reader_es = easyocr.Reader(['es'], gpu=True)  
device = torch.device("mps" if torch.backends.mps.is_available() else "cuda" if torch.cuda.is_available() else "cpu")
model_version="microsoft/trocr-small-printed"
processor=TrOCRProcessor.from_pretrained(model_version)
modelocr=VisionEncoderDecoderModel.from_pretrained(model_version).to(device)

def sistema_de_lectura():
    # Obtener la lista de archivos con extensión .PDF
    pdf_files_upper = glob.glob("PDFs_file/*.PDF")

    # Obtener la lista de archivos con extensión .pdf
    pdf_files_lower = glob.glob("PDFs_file/*.pdf")

    # Combinar ambas listas
    pdf_files = pdf_files_upper + pdf_files_lower


    for pdf_file in pdf_files:                                                  # Procesamos cada archivo PDF
        images = convert_from_path(pdf_file)                                    # y convertimos el PDF a imágenes
        for i, img in enumerate(images, start=1):                               # comenzando desde la imagen 1
            nombre_archivo = os.path.join(ruta_destino, f"factura{i}.jpg")      # hasta la ultima imagen 
            img.save(nombre_archivo)                                            # guardandolas en "img_of_pdf"
        
        # Esta sección reconoce los campos etiquetados y los guarda en un vector
        
        datos_totales = []      #Los datos extraidos se guardan en este vector
        
        def extract_number(filename):
            # Usa una expresión regular para encontrar números en el nombre del archivo
            match = re.search(r'\d+', filename)
            return int(match.group()) if match else float('inf')  # Retorna inf si no hay número para que estos archivos vayan al final

        # Obtén una lista de todas las imágenes en el directorio, ordenadas numéricamente
        image_paths = sorted(glob.glob(os.path.join(ruta_destino, "*.jpg")), key=lambda x: extract_number(os.path.basename(x)))

# Iterar sobre todas las imágenes en el directorio que contiene las imágenes
        for image_path in image_paths:
            
            results = model.predict(source=image_path, conf=0.5, save=False, line_width=2, show_labels=False, device='mps', imgsz=864)      # Detecta los campos de la imagen

            # Iterar sobre los resultados de la detección
            for result in results:
                boxes = result.boxes.cpu().numpy()  # Obtenemos las coordenadas de los campos reconocidos
                coords = boxes.xywhn                # coordenadas en el formato xywhn
                clases = boxes.cls                  # y nombres de los campos reconocidos

            image_easyocr = cv2.imread(image_path)
            height, width, _ = image_easyocr.shape

            textos_por_clase = [""] * 75

            # Iterar sobre un respectivo campo reconocido
            for coord, clase in zip(coords, clases):
                
                # Convertir coordenadas normalizadas a coordenadas absolutas
                x_center = float(coord[0]) * width
                y_center = float(coord[1]) * height
                w = float(coord[2]) * width
                h = float(coord[3]) * height

                # Calcular las coordenadas de las esquinas del rectángulo
                x = int(x_center - w / 2)
                y = int(y_center - h / 2)
                x2 = int(x_center + w / 2)
                y2 = int(y_center + h / 2)

                corrections = {
                r"\b s\b|\b1ss\b|\bC7 \b|\bVn 0\b|\bNIT\b|\bSn\b|\b~\b|\bs\b|\b;\b|\b =\b|\b: \b|\b:\b|\b. \b": "",
                r"=|~|\*": "",
                r'"': '',
                "=":"",
                "EETG |EETG.|EETG|EETC|EETC |E. ET.C. ": "EETC ",
                "CONSTRUCGION": "CONSTRUCCION",
                "GI-": "G1-",
                "LAPAZ": "LA PAZ",
                "ELALTO": "EL ALTO",
                "-GID-|-GDD-":"-GD-",
                "-ABP-|-AEF-|-AER-":"-ABR-",
                "-OCI-|-OCL-":"-OCT-",
                "-AG0-":"-AGO-",
                "TELEFEERICO|TELEFEFRICO|TELEFERRICO|TELEFEERICO":"TELEFERICO",
                "ABFIL|ABRL|ABRII|ABFL|ABELL|AERIL":"ABRIL",
                "JUNIC":"JUNIO",
                "JULIC|JULI0":"JULIO",
                "SEPTIERBRE|SEPTIMEMBRE|SEPTIERMBRE|SEPTIEMBERE|SEPTEMBERE|SEPTEMBRE|SEPTIMBRE":"SEPTIEMBRE",
                "DICIERBRE|DICLEMBRE|DICIERMBRE|DICIMBRE|DICIERNBRE|DICIERHBRE|DICEMBRE":"DICIEMBRE",
                "NOVIERBRE|NOVIMBRE|NOVIEWBRE":"NOVIEMBRE",
                "ESTAGION":"ESTACION",
                "UCELESTE":"L/CELESTE",
                "UAMARILLA,LAMARAILLA,UJAMARILLA,LJAMARILLA":"L/AMARILLA",
                "2800048029":"280048029",
                "985888-1-5":"985688-1-5",
                "985888-5":"985688-5",
                "LECTURA NORMAT|LECTURA NORMAI|LECTURA NOMMAL|LECTURA NOMAL":"LECTURA NORMAL",
                "LECLURA":"LECTURA",
                r"\bO KWH\b|\bO KWN\b|\bO KVARH\b|\bO KVARN\b|\bO KVAFN\b|\bO KW\b": "0",
                r"\bOKWH\b|\bOKWN\b|\bOKVARH\b|\bOKVARN\b|\bOKVAFN\b|\bOKW\b": "0",
                " KWH| KWN| KVARH| KVARN| KVAFN| KW|POTENCIA |PATENCIA |POLENCIA |POTENCLA ":"",
                "KWH|KWN|KVARH|KVARN|KVAFN|KW|POTENCIA|PATENCIA|POLENCIA|POTENCLA":"",
                "ENTE ":"ENFE "
                }

                if clase==0 or clase==15:
                    # Recortar la región de la imagen
                    roi_image = image_easyocr[y:y2, x:x2]

                    # Redimensionamiento
                    (height1, width1, canal) = roi_image.shape
                    redim_roi_image = cv2.resize(roi_image, (width1 * 7, height1 * 7))

                    # Aplicar reducción de ruido (noise reduction)
                    denoised_image = cv2.GaussianBlur(redim_roi_image, (5, 5), 0)

                    # Aumento de nitidez
                    kernel_sharpening = np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]])
                    sharpened_image = cv2.filter2D(denoised_image, -1, kernel_sharpening)

                    # Binarización
                    gray_image = cv2.cvtColor(sharpened_image, cv2.COLOR_BGR2GRAY)
                    _, binary_image = cv2.threshold(gray_image, 150, 255, cv2.THRESH_TOZERO)

                    valor = reader_es.readtext(binary_image, detail=0, paragraph=True)

                    for text in valor:
                        for pattern, replacement in corrections.items():
                            text = re.sub(pattern, replacement, text)
                    valor=text
                    
                else:
                    image = Image.open(image_path).convert("RGB")
                    crp_image=image.crop((x,y,x2,y2))
                    
                    pixel_values=processor(crp_image, return_tensors="pt").pixel_values.to(device)
                    generated_ids = modelocr.generate(pixel_values, max_new_tokens=50)
                    text_easy_ocr=processor.batch_decode(generated_ids, skip_special_tokens=True)[0]

                    for pattern, replacement in corrections.items():
                        text_easy_ocr = re.sub(pattern, replacement, text_easy_ocr)

                    texto_sin_puntos = text_easy_ocr.replace(",", "").replace(".", "").replace(" - ","").replace(" -","").replace("- ","").replace("-","")
                    if texto_sin_puntos.isnumeric():
                        # Consideramos que el último punto o coma es el separador decimal si existe
                        if ',' in text_easy_ocr or '.' in text_easy_ocr:
                            if text_easy_ocr.count('.') > 1:  # Si hay más de un punto
                                partes = text_easy_ocr.split('.')
                                texto_sin_comas = ''.join(partes[:-1]) + '.' + partes[-1]
                            elif text_easy_ocr.count('.') == 0 and text_easy_ocr.count(',') == 1:
                                texto_sin_comas = text_easy_ocr.replace(",", ".")
                            else:
                                texto_sin_comas = text_easy_ocr.replace(",", "")  # Elimina todas las comas
                            try:
                                valor = float(texto_sin_comas)
                            except ValueError:
                                valor = text_easy_ocr
                        elif '-' in text_easy_ocr:
                            valor=text_easy_ocr.replace(" - ","-").replace(" -","-").replace("- ","-").replace("-","-")
                        else:
                            valor = int(text_easy_ocr)
                    else:
                        valor = text_easy_ocr

                # Lista de mapeo que define cómo los valores de clase se corresponden a las posiciones en textos_por_clase
                mapeo_clase_a_posicion = [9, 19, 18, 20, 25, 24, 26, 22, 21, 23, 10, 13, 8, 4, 69, 7, 30, 42, 31, 29, 41, 43, 1, 15, 14, 74, 67, 44, 51, 45, 46, 48, 47, 50, 49, 53, 72, 61, 52, 58, 17, 16, 55, 56, 62, 68, 54, 6, 57, 32, 70, 65, 12, 28, 3, 71, 66, 33, 36, 37, 39, 38, 40, 11, 73, 2, 59, 60, 27, 34, 35, 63, 64, 5]

                # Asigna el valor a la posición correspondiente en textos_por_clase usando el mapeo
                textos_por_clase[mapeo_clase_a_posicion[int(clase)]] = valor
                
            textos_por_clase[0] = re.sub('img_of_pdf/|.jpg', '',re.sub('PDFs_file/','',pdf_file) + "_" + image_path )
            datos_totales.append(textos_por_clase)

        # Encontrar la primera fila vacía
        fila_inicio = hoja.max_row + 1

        # Escribir los datos en la hoja
        for fila, datos_imagen in enumerate(datos_totales, start=fila_inicio):
            for columna, dato_clase in enumerate(datos_imagen, start=1):
                if isinstance(dato_clase, list):
                    valor = ', '.join(map(str, dato_clase))  # Convertir la lista a una cadena separada por comas
                else:
                    valor = dato_clase
                celda = hoja.cell(row=fila, column=columna)
                celda.value = valor

        # Eliminar las imágenes procesadas
        for image_path in glob.glob(os.path.join(ruta_destino, "*.jpg")):
            os.remove(image_path)

    # Guardar el libro de trabajo en un archivo
    hoja.freeze_panes = hoja['B2']  #Inmoviliza paneles
    libro.save(file_path)
    libro.close()
    for pdf_file in pdf_files:
        os.remove(pdf_file)

    # Actualizar el mensaje con la ruta relativa
    current_dir = os.getcwd()
    relative_path = os.path.relpath(file_path, current_dir)
    mensaje_datos_extraidos.config(text=f"Los datos de las facturas han sido extraidos correctamente en:")
    mensaje_datos_extraidos2.config(text=f"{relative_path}")
    
    pass

def iniciar_lectura():
    mensaje.config(text="Espere mientras se extraen los datos...")
    mensaje_datos_extraidos.config(text="")
    mensaje_datos_extraidos2.config(text="")
    ventana.update()  # Actualiza la interfaz gráfica para mostrar el mensaje
    sistema_de_lectura()
    mensaje.config(text="")
##########################################FIN del sistema##########################################

import tkinter as tk
from tkinter import filedialog
import shutil
import os
from PIL import Image, ImageTk
def cargar_archivo():
    archivo = filedialog.askopenfilename()
    if archivo:
        guardar_en_carpeta(archivo)

def guardar_en_carpeta(ruta_archivo):
    carpeta_destino = "PDFs_file"  # Cambia esta ruta por la ruta de tu carpeta de destino
    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)
    
    nombre_archivo = os.path.basename(ruta_archivo)
    ruta_destino = os.path.join(carpeta_destino, nombre_archivo)
    
    try:
        shutil.copy(ruta_archivo, ruta_destino)
        print(f"Archivo copiado en: {ruta_destino}")
        mensaje_cargar.config(text=f"{nombre_archivo} cargado correctamente")

    except Exception as e:
        print(f"Error al copiar el archivo: {e}")
        mensaje_cargar.config(text=f"Error al cargar el archivo: {e}")
    

def limpiar_carpetas():
    carpetas = ["img_of_pdf", "PDFs_file"]
    for carpeta in carpetas:
        if os.path.exists(carpeta):
            for archivo in os.listdir(carpeta):
                archivo_path = os.path.join(carpeta, archivo)
                try:
                    if os.path.isfile(archivo_path) or os.path.islink(archivo_path):
                        os.unlink(archivo_path)
                    elif os.path.isdir(archivo_path):
                        shutil.rmtree(archivo_path)
                except Exception as e:
                    print(f'Error al borrar {archivo_path}. Razón: {e}')

def on_closing():
    limpiar_carpetas()
    ventana.destroy()   
####################INICIO DE LA INTERFAZ DE LA VENTANA PRINCIPAL##################################
# Crear la ventana principal
ventana = tk.Tk()
ventana.title("LECTOR DE FACTURAS ELECTRICAS")

# Obtener las dimensiones de la pantalla
ancho_pantalla = ventana.winfo_screenwidth()
alto_pantalla = ventana.winfo_screenheight()

# Calcular las coordenadas para centrar la ventana
ancho_ventana = 800
alto_ventana = 600
posicion_x = (ancho_pantalla - ancho_ventana) // 2 
posicion_y = (alto_pantalla - alto_ventana) // 2

# Establecer el tamaño y la posición de la ventana
ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{posicion_x}+{posicion_y}")

# Crear un Canvas
canvas = tk.Canvas(ventana, width=ancho_ventana, height=alto_ventana)
canvas.pack(fill="both", expand=True)

# Cargar la imagen de fondo
imagen_fondo = Image.open("/Users/jurgenalejandrorocasalvosanchez/Documents/Programa_PDG/Lector_de_Facturas/background.png")
imagen_fondo = imagen_fondo.resize((ancho_ventana, alto_ventana), Image.Resampling.LANCZOS)
imagen_fondo = ImageTk.PhotoImage(imagen_fondo)

# Dibujar la imagen de fondo
canvas.create_image(0, 0, anchor="nw", image=imagen_fondo)
fuente_texto=("Helvetica", 16, "bold")
# Crear y posicionar los widgets en el Canvas
etiqueta = tk.Label(ventana, text="Bienvenido al lector automatizado de facturas del servicio electrico", font=fuente_texto, bg="black")
canvas.create_window(ancho_ventana//2, 50, window=etiqueta)

btn_seleccionar = tk.Button(ventana, text="Cargue su PDF", font=fuente_texto, command=cargar_archivo)
canvas.create_window(ancho_ventana//2, 150, window=btn_seleccionar)

mensaje_cargar = tk.Label(ventana, text="", font=fuente_texto, bg="black", fg="white")
canvas.create_window(ancho_ventana//2, 200, window=mensaje_cargar)

mensaje = tk.Label(ventana, text="", font=fuente_texto, bg="black")
canvas.create_window(ancho_ventana//2, 250, window=mensaje)

mensaje_datos_extraidos = tk.Label(ventana, text="", font=fuente_texto, bg="black", fg="white")
canvas.create_window(ancho_ventana//2, 300, window=mensaje_datos_extraidos)
mensaje_datos_extraidos2 = tk.Label(ventana, text="", font=fuente_texto, bg="black", fg="white")
canvas.create_window(ancho_ventana//2, 350, window=mensaje_datos_extraidos2)

btn_iniciar = tk.Button(ventana, text="Inicie el programa", font=fuente_texto, command=iniciar_lectura)
canvas.create_window(ancho_ventana//2, 400, window=btn_iniciar)

etiqueta_desarrollador = tk.Label(ventana, text="Desarrollado por JaRs", font=fuente_texto, bg="black")
canvas.create_window(ancho_ventana//2, 550, window=etiqueta_desarrollador)

ventana.protocol("WM_DELETE_WINDOW", on_closing)

# Ejecutar el bucle principal de tkinter
ventana.mainloop()
###########################FIN de la interfaz##########################################



